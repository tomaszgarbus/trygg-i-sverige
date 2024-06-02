from openpyxl import load_workbook, Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import re
import pandas as pd
from functools import cache

_FNAME = '2023_Kommunala_resultat_NTU_2017-2022.xlsx'


def _read_table_descriptions(wb: Workbook) -> dict[str, str]:
    """Reads table descriptions and creates a mapping table_name->desc."""
    sheet = wb['Innehållsförteckning']
    descriptions = {}
    for row in range(7, 67):
        cell = f'C{row}'
        val = sheet[cell].value
        if type(val) == str and val.startswith('Tabell'):
            matches = re.match('Tabell (\d\.\d*) (.*)', val)
            if matches:
                descriptions[matches.groups()[0]] = matches.groups()[1]
    return descriptions


def _read_single_table_to_df(wb: Workbook, table_id: str) -> pd.DataFrame:
    """Reads a single Tabell to a df, without any manipulation.
    
    Suitable for tables 2.x and 3.x which follow the same formats and occupy
    the same Excel cells.
    Args:
        @table_id: in format "x.x" or "x.xx" without the word "Tabell"
    """
    # start = 'B5'
    # end = 'G338'
    ws = wb[f'Tabell {table_id}']
    values = list(map(lambda x: x[1:], list(ws.values)[4:338]))
    idx = list(map(lambda x: x[0], values[1:]))
    return pd.DataFrame(
        [v[1:] for v in values[1:]], columns=values[0][1:], index=idx)


def _build_df_for_year(
        wb: Workbook, year: str, table_series: int,
        numeric: bool = True) -> pd.DataFrame:
    """Builds a df with kommuns/läns as rows and crimes as columns.

    Ignores confidence intervals.
    Args:
        @year: Year in the "yyyy-yyyy" format as in original spreadsheet.
        @table_series: Either 2 or 3.
        @numeric: If true
    """
    assert year in ["2016-2017", "2018-2019", "2020-2021", "2021-2022"]
    year = year.replace("-", "–")
    assert table_series in [2, 3]
    table_ids = [k for k in _read_table_descriptions(wb).keys()
              if k[0] == str(table_series)]
    result = None
    for table_id in table_ids:
        table = _read_single_table_to_df(wb, table_id)
        table = table[year].to_frame()
        table = table.rename(columns={year: _read_table_descriptions(wb)[table_id]})
        if result is None:
            result = table
        else:
            result = result.join(table)
    if numeric:
        result = result.apply(lambda x: pd.to_numeric(x, errors='coerce'))
    return result


def default_crime_exposure_table() -> pd.DataFrame:
    """Default crime exposure table for last year."""
    return _build_df_for_year(
        load_workbook(_FNAME),
        "2020-2021",
        2
    )


def default_crime_fear_table() -> pd.DataFrame:
    """Default crime fear table for last year."""
    return _build_df_for_year(
        load_workbook(_FNAME),
        "2021-2022",
        3
    )


if __name__ == '__main__':
    # wb = load_workbook(_FNAME)
    # print(wb.sheetnames)
    # print(_read_table_descriptions(wb))
    # print(_read_single_table_to_df(wb, "2.12"))
    # print(_build_df_for_year(wb, "2020-2021", 2))
    pass
